#!/usr/bin/env python3
"""
Error-analysis taxonomy of false discoveries (#8).

The expert form labels every indicator as Genuine gap / Trivial / Illogical /
Already addressed. The non-genuine labels are *false discoveries*; this tool
turns them into a quantitative taxonomy, cross-tabulated with the Cooper
indicator type and the detection method, to answer "what KINDS of mistakes does
the system make, and which components produce them?" — a strong qualitative
chapter that complements the FDR number.

Usage (from backend/):
    python experiments/error_taxonomy.py \
        --forms experiments/expert_eval/expert_form_filled.xlsx \
        --results experiments/results/experiment_full_llama3.2_latest.json
"""

import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parent.parent
RESULTS_DIR = BACKEND_DIR / "experiments" / "results"

FALSE_LABELS = {"trivial", "illogical", "already addressed", "already-addressed"}


def _norm_type(raw: str) -> str:
    up = str(raw).upper()
    for t in ("FRAGMENTATION", "INCONSISTENCY", "INCOMPLETENESS"):
        if t in up:
            return t
    return "UNKNOWN"


def load_results_method_map(results_path):
    """description[:80] -> detection_method, from an experiment result JSON."""
    mapping = {}
    if not results_path or not Path(results_path).exists():
        return mapping
    report = json.loads(Path(results_path).read_text())
    for topic in report.get("phase3_gap_detection", {}).get("topics", []):
        for ind in topic.get("indicators", []):
            desc = (ind.get("description") or "")[:80]
            if desc:
                mapping[desc] = ind.get("detection_method") or "unspecified"
    return mapping


def read_form(form_path, method_map):
    """Return list of {label, type, method} for each evaluated indicator."""
    from openpyxl import load_workbook
    wb = load_workbook(form_path, data_only=True)
    ws = wb["Penilaian"] if "Penilaian" in wb.sheetnames else wb.active
    headers = {str(c.value).strip().lower(): i for i, c in enumerate(ws[1]) if c.value}
    li = headers.get("label")
    ti = headers.get("type")
    di = headers.get("description")
    rows = []
    if li is None:
        return rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        if li >= len(row) or row[li] is None:
            continue
        label = str(row[li]).strip().lower()
        ind_type = _norm_type(row[ti]) if ti is not None and row[ti] else "UNKNOWN"
        desc = str(row[di])[:80] if di is not None and row[di] else ""
        method = method_map.get(desc, "unspecified")
        rows.append({"label": label, "type": ind_type, "method": method})
    return rows


def render(rows) -> str:
    total = len(rows)
    false_rows = [r for r in rows if r["label"] in FALSE_LABELS]
    n_false = len(false_rows)
    genuine = sum(1 for r in rows if r["label"].startswith("genuine"))

    lines = [
        "# Taksonomi Error — Analisis False Discovery",
        f"\nTotal indikator dinilai: **{total}** · Genuine: **{genuine}** · "
        f"False discovery: **{n_false}** (FDR = {round(100 * n_false / total, 1) if total else 0}%)\n",
    ]

    # 1) Taxonomy by error category
    cat = Counter(r["label"] for r in false_rows)
    lines += ["## Kategori Error", "", "| Kategori | Jumlah | % dari false |",
              "|---|---|---|"]
    for label, n in cat.most_common():
        pct = round(100 * n / n_false, 1) if n_false else 0
        lines.append(f"| {label} | {n} | {pct}% |")
    lines.append("")

    # 2) Error category x indicator type
    by_type = defaultdict(Counter)
    for r in false_rows:
        by_type[r["type"]][r["label"]] += 1
    cats = [c for c, _ in cat.most_common()]
    lines += ["## Error × Tipe Indikator", "",
              "| Tipe | " + " | ".join(cats) + " | Total |",
              "|---|" + "---|" * (len(cats) + 1)]
    for t in ("FRAGMENTATION", "INCONSISTENCY", "INCOMPLETENESS", "UNKNOWN"):
        if t not in by_type:
            continue
        counts = [str(by_type[t].get(c, 0)) for c in cats]
        lines.append(f"| {t} | " + " | ".join(counts) + f" | {sum(by_type[t].values())} |")
    lines.append("")

    # 3) Error category x detection method
    by_method = defaultdict(Counter)
    for r in false_rows:
        by_method[r["method"]][r["label"]] += 1
    lines += ["## Error × Metode Deteksi", "",
              "| Metode | " + " | ".join(cats) + " | Total |",
              "|---|" + "---|" * (len(cats) + 1)]
    for m, c in sorted(by_method.items(), key=lambda kv: -sum(kv[1].values())):
        counts = [str(c.get(cc, 0)) for cc in cats]
        lines.append(f"| `{m}` | " + " | ".join(counts) + f" | {sum(c.values())} |")
    lines += ["", "_Menunjukkan jenis kesalahan dominan dan komponen/metode "
              "yang paling sering menghasilkannya — dasar perbaikan terarah._"]
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="False-discovery error taxonomy from expert form.")
    parser.add_argument("--forms", action="append", required=True)
    parser.add_argument("--results", default=None, help="Result JSON (to map detection_method)")
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    method_map = load_results_method_map(args.results)
    rows = []
    for f in args.forms:
        p = Path(f)
        if p.exists():
            rows.extend(read_form(p, method_map))
        else:
            print(f"  ! form not found: {p}")

    if not rows:
        print("No labeled rows found (need a filled 'label' column).")
        return 1

    md = render(rows)
    out = Path(args.output) if args.output else RESULTS_DIR / "error_taxonomy.md"
    out.write_text(md)
    print(f"Taksonomi tersimpan: {out}\n")
    print(md)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
