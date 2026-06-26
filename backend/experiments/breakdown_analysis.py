#!/usr/bin/env python3
"""
Breakdown analysis — decompose gap-detection results by indicator type and by
detection method.

The thesis claims a neuro-symbolic *system*; this report shows WHICH parts carry
the load: how indicators (and their post-rule-engine acceptance) distribute
across the three Cooper indicator types (fragmentation / inconsistency /
incompleteness) and across detection methods (topic_clustering, fact_table_
contradicts, nli_cross_encoder, llm_nli, aspect_coverage, methodology_coverage).

When an expert-evaluation form is provided, it also breaks down the Expert
Acceptance Rate (EAR) and mean Logical Coherence Score (LCS) per type/method,
so you can argue e.g. "the dedicated NLI signal yields higher-acceptance
inconsistency indicators than the LLM-only path."

Usage (from backend/):
    python experiments/breakdown_analysis.py \
        --results experiments/results/experiment_full_llama3.2_latest.json
    # optional expert form to add EAR/LCS columns:
    python experiments/breakdown_analysis.py --results <json> \
        --expert-form experiments/expert_eval/expert_form_filled.xlsx
"""

import argparse
import json
import statistics
from collections import defaultdict
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parent.parent
RESULTS_DIR = BACKEND_DIR / "experiments" / "results"


def load_indicators(results_path: Path):
    """Flatten all indicators across topics into a single list of dicts."""
    report = json.loads(results_path.read_text())
    indicators = []
    for topic in report.get("phase3_gap_detection", {}).get("topics", []):
        topic_name = topic.get("topic", topic.get("topic_key", "?"))
        for ind in topic.get("indicators", []):
            indicators.append({
                "topic": topic_name,
                "type": _norm_type(ind.get("type", "UNKNOWN")),
                "method": ind.get("detection_method") or "unspecified",
                "confidence": float(ind.get("confidence", 0.0) or 0.0),
                "verdict": _norm_verdict(ind.get("verdict", "NONE")),
                "description": ind.get("description", ""),
            })
    return indicators, report


def _norm_type(raw: str) -> str:
    up = str(raw).upper()
    for t in ("FRAGMENTATION", "INCONSISTENCY", "INCOMPLETENESS"):
        if t in up:
            return t
    return "UNKNOWN"


def _norm_verdict(raw: str) -> str:
    up = str(raw).upper()
    for v in ("PASS", "FLAG", "REJECT"):
        if v in up:
            return v
    return "NONE"


def _group_stats(rows, key):
    """Aggregate count / mean confidence / verdict mix per group key."""
    groups = defaultdict(list)
    for r in rows:
        groups[r[key]].append(r)
    out = []
    for name, items in sorted(groups.items(), key=lambda kv: -len(kv[1])):
        confs = [i["confidence"] for i in items]
        rejected = sum(1 for i in items if i["verdict"] == "REJECT")
        flagged = sum(1 for i in items if i["verdict"] == "FLAG")
        passed = sum(1 for i in items if i["verdict"] == "PASS")
        out.append({
            "name": name,
            "count": len(items),
            "mean_conf": round(statistics.mean(confs), 3) if confs else 0.0,
            "pass": passed, "flag": flagged, "reject": rejected,
        })
    return out


def _load_expert(form_path: Path):
    """Map indicator description -> {genuine: bool, lcs: float} from an expert form."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  (openpyxl not available — skipping expert breakdown)")
        return {}
    wb = load_workbook(form_path, data_only=True)
    ws = wb["Penilaian"] if "Penilaian" in wb.sheetnames else wb.active
    headers = {str(c.value).strip().lower(): i for i, c in enumerate(ws[1]) if c.value}
    desc_i = headers.get("description") or headers.get("deskripsi")
    label_i = headers.get("label")
    lcs_i = headers.get("lcs_1to5") or headers.get("lcs")
    mapping = {}
    if desc_i is None:
        return mapping
    for row in ws.iter_rows(min_row=2, values_only=True):
        desc = str(row[desc_i])[:80] if row[desc_i] else None
        if not desc:
            continue
        label = str(row[label_i]).strip().lower() if label_i is not None and row[label_i] else ""
        lcs = None
        if lcs_i is not None and row[lcs_i] is not None:
            try:
                lcs = float(row[lcs_i])
            except (ValueError, TypeError):
                lcs = None
        mapping[desc] = {"genuine": label.startswith("genuine"), "lcs": lcs}
    return mapping


def _expert_stats(rows, key, expert_map):
    groups = defaultdict(list)
    for r in rows:
        ev = expert_map.get(r["description"][:80])
        if ev is not None:
            groups[r[key]].append(ev)
    out = []
    for name, evs in sorted(groups.items(), key=lambda kv: -len(kv[1])):
        n = len(evs)
        ear = 100.0 * sum(1 for e in evs if e["genuine"]) / n if n else 0.0
        lcs_vals = [e["lcs"] for e in evs if e["lcs"] is not None]
        lcs = round(statistics.mean(lcs_vals), 2) if lcs_vals else None
        out.append({"name": name, "n": n, "ear": round(ear, 1), "lcs": lcs})
    return out


def render(results_path: Path, expert_form: Path | None) -> str:
    rows, report = load_indicators(results_path)
    info = report.get("experiment_info", {})
    lines = [
        f"# Breakdown Analysis — {results_path.name}",
        f"\nMode: `{info.get('mode', '?')}` · Model: `{info.get('model', info.get('model_name', '?'))}` "
        f"· Total indikator: {len(rows)}\n",
    ]

    for key, title in [("type", "Tipe Indikator (Cooper)"), ("method", "Metode Deteksi")]:
        lines += [f"## Per {title}", "",
                  "| Grup | Jumlah | Mean Conf | PASS | FLAG | REJECT |",
                  "|---|---|---|---|---|---|"]
        for g in _group_stats(rows, key):
            lines.append(f"| `{g['name']}` | {g['count']} | {g['mean_conf']} "
                         f"| {g['pass']} | {g['flag']} | {g['reject']} |")
        lines.append("")

    if expert_form and expert_form.exists():
        expert_map = _load_expert(expert_form)
        if expert_map:
            for key, title in [("type", "Tipe Indikator"), ("method", "Metode Deteksi")]:
                lines += [f"## Evaluasi Pakar per {title} (EAR / LCS)", "",
                          "| Grup | n dinilai | EAR % | LCS rata² |",
                          "|---|---|---|---|"]
                for g in _expert_stats(rows, key, expert_map):
                    lines.append(f"| `{g['name']}` | {g['n']} | {g['ear']} "
                                 f"| {g['lcs'] if g['lcs'] is not None else '—'} |")
                lines.append("")

    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Breakdown by indicator type & detection method.")
    parser.add_argument("--results", required=True, help="Experiment results JSON")
    parser.add_argument("--expert-form", default=None, help="Optional filled expert form (XLSX)")
    parser.add_argument("--output", default=None, help="Output Markdown path")
    args = parser.parse_args()

    results_path = Path(args.results)
    if not results_path.is_absolute() and not results_path.exists():
        results_path = RESULTS_DIR / results_path.name
    expert_form = Path(args.expert_form) if args.expert_form else None

    md = render(results_path, expert_form)
    out = Path(args.output) if args.output else results_path.with_name(
        results_path.stem + "_breakdown.md")
    out.write_text(md)
    print(f"Breakdown tersimpan: {out}\n")
    print(md)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
