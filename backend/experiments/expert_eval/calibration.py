#!/usr/bin/env python3
"""
Confidence calibration analysis for the gap-indicator confidence scores.

A well-calibrated system whose indicators carry confidence 0.8 should have ~80%
of those indicators judged genuine by experts. This script turns the Priority-1
evidence-based confidence calibration into a *measurable* claim by computing,
from a filled expert form:

    - Brier score      mean((confidence - genuine)^2), lower is better
    - ECE              Expected Calibration Error (binned |confidence - accuracy|)
    - MCE              Maximum Calibration Error
    - Reliability table per confidence bin (mean conf vs observed genuine rate)

Outcome label: an indicator counts as a positive (genuine = 1) when the expert
label is "Genuine gap", else 0. Predicted probability = the system confidence.

Usage (from backend/):
    python experiments/expert_eval/calibration.py \
        --forms experiments/expert_eval/expert_form_filled.xlsx
    # compare raw vs rule-engine-adjusted confidence:
    python experiments/expert_eval/calibration.py --forms <xlsx> --use-adjusted
"""

import argparse
import json
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parents[2]


def _read_pairs(form_path: Path, use_adjusted: bool):
    """Return list of (confidence, genuine:int) from a filled expert form."""
    from openpyxl import load_workbook
    wb = load_workbook(form_path, data_only=True)
    ws = wb["Penilaian"] if "Penilaian" in wb.sheetnames else wb.active
    headers = {str(c.value).strip().lower(): i for i, c in enumerate(ws[1]) if c.value}
    conf_key = "adjusted_confidence" if use_adjusted else "confidence"
    ci = headers.get(conf_key)
    if ci is None:
        ci = headers.get("confidence")
    li = headers.get("label")
    pairs = []
    if ci is None or li is None:
        return pairs
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[ci] is None or row[li] is None:
            continue
        try:
            conf = float(row[ci])
        except (ValueError, TypeError):
            continue
        genuine = 1 if str(row[li]).strip().lower().startswith("genuine") else 0
        pairs.append((conf, genuine))
    return pairs


def calibration_metrics(pairs, n_bins: int = 10):
    """Compute Brier, ECE, MCE, and per-bin reliability stats."""
    n = len(pairs)
    if n == 0:
        return None
    brier = sum((c - y) ** 2 for c, y in pairs) / n

    bins = [[] for _ in range(n_bins)]
    for c, y in pairs:
        idx = min(int(c * n_bins), n_bins - 1)
        bins[idx].append((c, y))

    ece, mce = 0.0, 0.0
    table = []
    for b, items in enumerate(bins):
        lo, hi = b / n_bins, (b + 1) / n_bins
        if not items:
            table.append({"bin": f"[{lo:.1f},{hi:.1f})", "n": 0,
                          "mean_conf": None, "genuine_rate": None, "gap": None})
            continue
        mean_conf = sum(c for c, _ in items) / len(items)
        acc = sum(y for _, y in items) / len(items)
        gap = abs(mean_conf - acc)
        ece += (len(items) / n) * gap
        mce = max(mce, gap)
        table.append({
            "bin": f"[{lo:.1f},{hi:.1f})", "n": len(items),
            "mean_conf": round(mean_conf, 3), "genuine_rate": round(acc, 3),
            "gap": round(gap, 3),
        })

    return {
        "n": n,
        "brier": round(brier, 4),
        "ece": round(ece, 4),
        "mce": round(mce, 4),
        "base_rate": round(sum(y for _, y in pairs) / n, 3),
        "reliability": table,
    }


def render(metrics, use_adjusted: bool) -> str:
    which = "adjusted_confidence (post Rule Engine)" if use_adjusted else "confidence (raw)"
    lines = [
        "# Kalibrasi Confidence — Indikator Synthesis Gap",
        f"\nProbabilitas prediksi: **{which}** · Outcome: label = *Genuine gap*",
        f"\n- Indikator dinilai: **{metrics['n']}**",
        f"- Base rate (genuine): **{metrics['base_rate']}**",
        f"- **Brier score**: {metrics['brier']} (semakin kecil semakin baik; "
        f"prediksi konstan = {round(metrics['base_rate'] * (1 - metrics['base_rate']), 4)})",
        f"- **ECE** (Expected Calibration Error): {metrics['ece']}",
        f"- **MCE** (Maximum Calibration Error): {metrics['mce']}",
        "",
        "## Reliability diagram (data)",
        "",
        "| Bin confidence | n | Mean conf | Genuine rate | |gap| |",
        "|---|---|---|---|---|",
    ]
    for r in metrics["reliability"]:
        if r["n"] == 0:
            lines.append(f"| {r['bin']} | 0 | — | — | — |")
        else:
            lines.append(f"| {r['bin']} | {r['n']} | {r['mean_conf']} "
                         f"| {r['genuine_rate']} | {r['gap']} |")
    lines += [
        "",
        "_Kalibrasi sempurna: mean conf ≈ genuine rate di tiap bin (gap≈0). "
        "ECE adalah rata-rata gap berbobot jumlah indikator per bin._",
    ]
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Confidence calibration (Brier/ECE) from expert forms.")
    parser.add_argument("--forms", action="append", required=True,
                        help="Filled expert form XLSX (repeatable to pool raters)")
    parser.add_argument("--use-adjusted", action="store_true",
                        help="Use adjusted_confidence (post Rule Engine) instead of raw confidence")
    parser.add_argument("--bins", type=int, default=10)
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    pairs = []
    for f in args.forms:
        p = Path(f)
        if not p.exists():
            print(f"  ! form not found: {p}")
            continue
        pairs.extend(_read_pairs(p, args.use_adjusted))

    metrics = calibration_metrics(pairs, n_bins=args.bins)
    if metrics is None:
        print("No (confidence, label) pairs found. Is the form filled (column 'label')?")
        return 1

    md = render(metrics, args.use_adjusted)
    base = Path(args.forms[0])
    out = Path(args.output) if args.output else base.with_name("calibration_metrics.md")
    out.write_text(md)
    out.with_suffix(".json").write_text(json.dumps(metrics, indent=2))
    print(f"Kalibrasi tersimpan: {out}\n")
    print(md)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
