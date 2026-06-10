#!/usr/bin/env python3
"""
Expert Evaluation Metrics Calculator — Wizard Research

Reads one or more filled expert assessment forms (XLSX, from
generate_form.py) and computes the thesis evaluation metrics
(RINGKASAN_REVISI.md §11):

    EAR   Expert Acceptance Rate     — % indicators judged "Genuine gap"
    FDR   False Discovery Rate       — % indicators judged NOT genuine
    LCS   Logical Coherence Score    — mean expert score (1-5)
    AS    Actionability Score        — mean expert score (1-5)
    SHG   Semantic vs Human Gap      — Spearman correlation between the
                                        system's confidence ranking and the
                                        expert's importance ranking (per topic)
    REP   Rule Engine Precision      — % of REJECT verdicts the expert
                                        confirms as justified

Hypothesis tests:
    H4: EAR >= 50%
    H5: LCS >= 3.5

With multiple raters, per-rater metrics plus Cohen's kappa (label agreement,
first two raters) are reported.

Usage:
    cd backend
    python experiments/expert_eval/compute_metrics.py \
        --forms experiments/expert_eval/expert_form_filled.xlsx \
        [--forms second_rater.xlsx] \
        [--output experiments/expert_eval/expert_metrics.json]
"""

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from scipy.stats import spearmanr

GENUINE = "genuine gap"
H4_THRESHOLD = 50.0   # EAR >= 50%
H5_THRESHOLD = 3.5    # LCS >= 3.5


def read_form(path: Path) -> list:
    """Read assessment rows from a filled XLSX form."""
    wb = load_workbook(path, data_only=True)
    if "Penilaian" not in wb.sheetnames:
        print(f"ERROR: sheet 'Penilaian' not found in {path}")
        sys.exit(1)
    ws = wb["Penilaian"]

    header = [str(c.value) if c.value is not None else "" for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    required = ["indicator_id", "topic_key", "system_verdict", "system_rank",
                "label", "lcs_1to5", "as_1to5", "expert_rank", "reject_justified"]
    missing = [c for c in required if c not in idx]
    if missing:
        print(f"ERROR: missing columns {missing} in {path}")
        sys.exit(1)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[idx["indicator_id"]] in (None, ""):
            continue
        rows.append({
            "indicator_id": str(row[idx["indicator_id"]]),
            "topic_key": str(row[idx["topic_key"]]),
            "system_verdict": str(row[idx["system_verdict"]] or "").upper(),
            "system_rank": row[idx["system_rank"]],
            "label": str(row[idx["label"]] or "").strip(),
            "lcs": row[idx["lcs_1to5"]],
            "as": row[idx["as_1to5"]],
            "expert_rank": row[idx["expert_rank"]],
            "reject_justified": str(row[idx["reject_justified"]] or "").strip().lower(),
        })
    return rows


def compute_rater_metrics(rows: list) -> dict:
    """Compute all metrics for a single rater's form."""
    labeled = [r for r in rows if r["label"]]
    n = len(labeled)
    if n == 0:
        return {"error": "no labeled rows — form appears unfilled"}

    genuine = sum(1 for r in labeled if r["label"].lower() == GENUINE)
    label_counts = defaultdict(int)
    for r in labeled:
        label_counts[r["label"]] += 1

    lcs_scores = [float(r["lcs"]) for r in labeled if isinstance(r["lcs"], (int, float))]
    as_scores = [float(r["as"]) for r in labeled if isinstance(r["as"], (int, float))]

    ear = genuine / n * 100
    fdr = 100 - ear
    lcs = sum(lcs_scores) / len(lcs_scores) if lcs_scores else None
    as_score = sum(as_scores) / len(as_scores) if as_scores else None

    # SHG: Spearman per topic between system_rank and expert_rank, then average
    shg_per_topic = {}
    by_topic = defaultdict(list)
    for r in labeled:
        if isinstance(r["system_rank"], (int, float)) and isinstance(r["expert_rank"], (int, float)):
            by_topic[r["topic_key"]].append((float(r["system_rank"]), float(r["expert_rank"])))
    for topic, pairs in by_topic.items():
        if len(pairs) >= 3:  # Spearman needs >= 3 points to be meaningful
            sys_ranks, exp_ranks = zip(*pairs)
            rho, pval = spearmanr(sys_ranks, exp_ranks)
            shg_per_topic[topic] = {"spearman_rho": round(float(rho), 3),
                                    "p_value": round(float(pval), 4),
                                    "n": len(pairs)}
    shg_values = [v["spearman_rho"] for v in shg_per_topic.values()]
    shg_avg = round(sum(shg_values) / len(shg_values), 3) if shg_values else None

    # REP: of system REJECTs, % the expert confirms as justified
    rejects = [r for r in rows if r["system_verdict"] == "REJECT"
               and r["reject_justified"] in ("yes", "no")]
    rep = (sum(1 for r in rejects if r["reject_justified"] == "yes") / len(rejects) * 100
           if rejects else None)

    return {
        "n_assessed": n,
        "label_distribution": dict(label_counts),
        "EAR_percent": round(ear, 1),
        "FDR_percent": round(fdr, 1),
        "LCS_mean": round(lcs, 2) if lcs is not None else None,
        "AS_mean": round(as_score, 2) if as_score is not None else None,
        "SHG_spearman_avg": shg_avg,
        "SHG_per_topic": shg_per_topic,
        "REP_percent": round(rep, 1) if rep is not None else None,
        "REP_n_rejects_assessed": len(rejects),
        "hypotheses": {
            "H4_EAR_geq_50": {
                "threshold": H4_THRESHOLD,
                "value": round(ear, 1),
                "supported": ear >= H4_THRESHOLD,
            },
            "H5_LCS_geq_3.5": {
                "threshold": H5_THRESHOLD,
                "value": round(lcs, 2) if lcs is not None else None,
                "supported": (lcs >= H5_THRESHOLD) if lcs is not None else None,
            },
        },
    }


def cohens_kappa(labels_a: list, labels_b: list) -> float:
    """Cohen's kappa for two raters' label lists (paired by index)."""
    assert len(labels_a) == len(labels_b)
    n = len(labels_a)
    if n == 0:
        return float("nan")
    categories = sorted(set(labels_a) | set(labels_b))
    po = sum(1 for a, b in zip(labels_a, labels_b) if a == b) / n
    pe = sum(
        (labels_a.count(c) / n) * (labels_b.count(c) / n)
        for c in categories
    )
    return (po - pe) / (1 - pe) if pe < 1 else 1.0


def to_markdown(report: dict) -> str:
    """Render the aggregate report as a Markdown table for BAB IV."""
    lines = ["| Metrik | Nilai | Keterangan |", "|---|---|---|"]
    agg = report["aggregate"]
    lines.append(f"| EAR (Expert Acceptance Rate) | {agg['EAR_percent']}% | "
                 f"H4 (≥50%): {'TERDUKUNG' if agg['hypotheses']['H4_EAR_geq_50']['supported'] else 'TIDAK terdukung'} |")
    lines.append(f"| FDR (False Discovery Rate) | {agg['FDR_percent']}% | komplemen EAR |")
    lcs_h5 = agg['hypotheses']['H5_LCS_geq_3.5']['supported']
    lines.append(f"| LCS (Logical Coherence Score) | {agg['LCS_mean']}/5 | "
                 f"H5 (≥3.5): {'TERDUKUNG' if lcs_h5 else 'TIDAK terdukung'} |")
    lines.append(f"| AS (Actionability Score) | {agg['AS_mean']}/5 | — |")
    lines.append(f"| SHG (Spearman sistem vs pakar) | {agg['SHG_spearman_avg']} | rata-rata antar topik |")
    rep = agg.get("REP_percent")
    lines.append(f"| REP (Rule Engine Precision) | {rep if rep is not None else 'N/A'}"
                 f"{'%' if rep is not None else ''} | dari verdict REJECT |")
    if report.get("inter_rater"):
        ir = report["inter_rater"]
        lines.append(f"| Cohen's κ (2 penilai pertama) | {ir['cohens_kappa']} | "
                     f"{ir['n_paired']} indikator |")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Compute expert evaluation metrics")
    parser.add_argument(
        "--forms", action="append", required=True,
        help="Filled XLSX form path (repeat for multiple raters)",
    )
    parser.add_argument(
        "--output", default=None,
        help="Output JSON path (default: <first_form_dir>/expert_metrics.json)",
    )
    args = parser.parse_args()

    form_paths = [Path(f) for f in args.forms]
    for p in form_paths:
        if not p.exists():
            print(f"ERROR: form not found: {p}")
            sys.exit(1)

    per_rater = {}
    all_rows = {}
    for i, path in enumerate(form_paths, start=1):
        rows = read_form(path)
        rater_id = f"rater_{i}"
        per_rater[rater_id] = {
            "form": str(path),
            "metrics": compute_rater_metrics(rows),
        }
        all_rows[rater_id] = rows

    # Aggregate = average across raters (or single rater passthrough)
    valid = [r["metrics"] for r in per_rater.values() if "error" not in r["metrics"]]
    if not valid:
        print("ERROR: no valid (filled) forms")
        sys.exit(1)

    def avg(key):
        vals = [m[key] for m in valid if m.get(key) is not None]
        return round(sum(vals) / len(vals), 2) if vals else None

    ear = avg("EAR_percent")
    lcs = avg("LCS_mean")
    aggregate = {
        "n_raters": len(valid),
        "EAR_percent": ear,
        "FDR_percent": avg("FDR_percent"),
        "LCS_mean": lcs,
        "AS_mean": avg("AS_mean"),
        "SHG_spearman_avg": avg("SHG_spearman_avg"),
        "REP_percent": avg("REP_percent"),
        "hypotheses": {
            "H4_EAR_geq_50": {
                "threshold": H4_THRESHOLD, "value": ear,
                "supported": (ear is not None and ear >= H4_THRESHOLD),
            },
            "H5_LCS_geq_3.5": {
                "threshold": H5_THRESHOLD, "value": lcs,
                "supported": (lcs is not None and lcs >= H5_THRESHOLD),
            },
        },
    }

    report = {"per_rater": per_rater, "aggregate": aggregate}

    # Inter-rater agreement (first two raters, paired by indicator_id)
    if len(all_rows) >= 2:
        raters = list(all_rows)
        rows_a = {r["indicator_id"]: r["label"] for r in all_rows[raters[0]] if r["label"]}
        rows_b = {r["indicator_id"]: r["label"] for r in all_rows[raters[1]] if r["label"]}
        common = sorted(set(rows_a) & set(rows_b))
        if common:
            kappa = cohens_kappa([rows_a[i] for i in common], [rows_b[i] for i in common])
            report["inter_rater"] = {
                "raters": raters[:2],
                "n_paired": len(common),
                "cohens_kappa": round(kappa, 3),
            }

    output_path = Path(args.output) if args.output else form_paths[0].parent / "expert_metrics.json"
    output_path.write_text(json.dumps(report, indent=2, ensure_ascii=False))

    print(f"Metrics saved: {output_path}\n")
    print("=== TABEL UNTUK BAB IV ===\n")
    print(to_markdown(report))


if __name__ == "__main__":
    main()
