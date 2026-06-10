#!/usr/bin/env python3
"""
Expert Evaluation Form Generator — Wizard Research

Reads gap indicators from experiment result JSON files and produces an XLSX
assessment form for domain experts. Expert judgments feed the thesis
evaluation metrics (EAR, LCS, AS, FDR, SHG, REP — see compute_metrics.py).

Per-indicator expert inputs (per RINGKASAN_REVISI.md §11):
    label        Genuine gap | Trivial | Illogical | Already addressed
    lcs          Logical Coherence Score, 1-5
    as           Actionability Score, 1-5
    expert_rank  Expert's importance ranking within each topic (1 = most important)
    reject_justified  (only for REJECT verdicts) yes | no
    notes        Free-text comments

Usage:
    cd backend
    python experiments/expert_eval/generate_form.py \
        --results experiments/results/experiment_full_llama3.2_latest.json \
        --output experiments/expert_eval/expert_form.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BACKEND_DIR = Path(__file__).resolve().parent.parent.parent
DEFAULT_RESULTS = BACKEND_DIR / "experiments" / "results"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "expert_form.xlsx"

LABELS = ["Genuine gap", "Trivial", "Illogical", "Already addressed"]
SCORE_RANGE = "1,2,3,4,5"

SYSTEM_COLUMNS = [
    ("indicator_id", 14),
    ("topic_key", 10),
    ("topic", 38),
    ("type", 18),
    ("description", 60),
    ("confidence", 11),
    ("adjusted_confidence", 11),
    ("system_verdict", 13),
    ("system_rank", 11),
    ("suggested_directions", 50),
]

EXPERT_COLUMNS = [
    ("label", 20),
    ("lcs_1to5", 10),
    ("as_1to5", 10),
    ("expert_rank", 12),
    ("reject_justified", 15),
    ("notes", 45),
]


def collect_indicators(results_path: Path) -> list:
    """Extract indicator rows from an experiment result JSON."""
    data = json.loads(results_path.read_text())
    mode = data.get("experiment_info", {}).get("mode", "full")
    model = data.get("experiment_info", {}).get("model", "unknown")

    rows = []
    for topic_data in data.get("phase3_gap_detection", {}).get("topics", []):
        if "error" in topic_data:
            continue
        topic_key = topic_data.get("topic_key", "?")
        topic = topic_data.get("topic", "")

        # System rank = order by confidence within topic (1 = highest)
        indicators = topic_data.get("indicators", [])
        ranked = sorted(
            enumerate(indicators),
            key=lambda x: x[1].get("confidence", 0),
            reverse=True,
        )
        rank_of = {idx: rank + 1 for rank, (idx, _) in enumerate(ranked)}

        for idx, ind in enumerate(indicators):
            rows.append({
                "indicator_id": f"{topic_key}-{idx + 1:02d}",
                "topic_key": topic_key,
                "topic": topic,
                "type": ind.get("type", ""),
                "description": ind.get("description", ""),
                "confidence": ind.get("confidence", 0),
                "adjusted_confidence": ind.get("adjusted_confidence", 0),
                "system_verdict": ind.get("verdict", "NONE"),
                "system_rank": rank_of[idx],
                "suggested_directions": "; ".join(ind.get("suggested_directions", []))[:300],
                "mode": mode,
                "model": model,
            })
    return rows


def build_workbook(rows: list, source_name: str) -> Workbook:
    wb = Workbook()

    # --- Sheet 1: instructions ---
    ws_info = wb.active
    ws_info.title = "Petunjuk"
    instructions = [
        ["FORM PENILAIAN PAKAR — Indikator Synthesis Gap"],
        [""],
        ["Sumber data", source_name],
        [""],
        ["Cara mengisi (sheet 'Penilaian'):"],
        ["1. label", "Genuine gap = indikator valid; Trivial = terlalu dangkal; "
                     "Illogical = cacat logika; Already addressed = sudah dijawab literatur"],
        ["2. lcs_1to5", "Logical Coherence Score: 1 = tidak logis ... 5 = sangat logis"],
        ["3. as_1to5", "Actionability Score: 1 = tidak dapat ditindaklanjuti ... 5 = sangat spesifik"],
        ["4. expert_rank", "Peringkat kepentingan indikator DALAM SATU TOPIK (1 = paling penting)"],
        ["5. reject_justified", "Hanya untuk baris dengan system_verdict = REJECT: apakah penolakan tepat? (yes/no)"],
        ["6. notes", "Komentar bebas (opsional)"],
        [""],
        ["Kolom sistem (indicator_id .. suggested_directions) JANGAN diubah."],
    ]
    for row in instructions:
        ws_info.append(row)
    ws_info.column_dimensions["A"].width = 18
    ws_info.column_dimensions["B"].width = 100
    ws_info["A1"].font = Font(bold=True, size=14)

    # --- Sheet 2: assessment ---
    ws = wb.create_sheet("Penilaian")
    header = [c for c, _ in SYSTEM_COLUMNS] + [c for c, _ in EXPERT_COLUMNS]
    ws.append(header)

    sys_fill = PatternFill("solid", fgColor="DDEBF7")   # blue: system data
    exp_fill = PatternFill("solid", fgColor="E2EFDA")   # green: expert input
    for col_idx, name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = sys_fill if col_idx <= len(SYSTEM_COLUMNS) else exp_fill
        width = ([w for _, w in SYSTEM_COLUMNS] + [w for _, w in EXPERT_COLUMNS])[col_idx - 1]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in rows:
        ws.append([row[c] for c, _ in SYSTEM_COLUMNS] + [""] * len(EXPERT_COLUMNS))

    n = len(rows)
    if n:
        first_expert_col = len(SYSTEM_COLUMNS) + 1

        dv_label = DataValidation(
            type="list", formula1='"' + ",".join(LABELS) + '"', allow_blank=True
        )
        dv_score = DataValidation(type="list", formula1='"' + SCORE_RANGE + '"', allow_blank=True)
        dv_yesno = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)

        for dv in (dv_label, dv_score, dv_yesno):
            ws.add_data_validation(dv)

        label_col = get_column_letter(first_expert_col)
        lcs_col = get_column_letter(first_expert_col + 1)
        as_col = get_column_letter(first_expert_col + 2)
        rj_col = get_column_letter(first_expert_col + 4)

        dv_label.add(f"{label_col}2:{label_col}{n + 1}")
        dv_score.add(f"{lcs_col}2:{lcs_col}{n + 1}")
        dv_score.add(f"{as_col}2:{as_col}{n + 1}")
        dv_yesno.add(f"{rj_col}2:{rj_col}{n + 1}")

        # Wrap long text columns
        for col_name in ("description", "topic", "suggested_directions"):
            col_idx = header.index(col_name) + 1
            for r in range(2, n + 2):
                ws.cell(row=r, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    return wb


def main():
    parser = argparse.ArgumentParser(description="Generate expert evaluation XLSX form")
    parser.add_argument(
        "--results", required=True,
        help="Path to experiment result JSON (e.g. experiments/results/experiment_full_llama3.2_latest.json)",
    )
    parser.add_argument(
        "--output", default=str(DEFAULT_OUTPUT),
        help=f"Output XLSX path (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()

    results_path = Path(args.results)
    if not results_path.exists():
        print(f"ERROR: results file not found: {results_path}")
        sys.exit(1)

    rows = collect_indicators(results_path)
    if not rows:
        print("ERROR: no indicators found in results file")
        sys.exit(1)

    wb = build_workbook(rows, results_path.name)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"Form generated : {output_path}")
    print(f"Indicators     : {len(rows)}")
    print(f"Topics         : {sorted({r['topic_key'] for r in rows})}")
    print(f"REJECT rows    : {sum(1 for r in rows if r['system_verdict'] == 'REJECT')}")


if __name__ == "__main__":
    main()
