#!/usr/bin/env python3
"""
SPO Fact Precision Annotation — Wizard Research

Measures extraction quality of the Fact Base (foundation of the Rule Engine).
Samples N facts from an experiment result, produces an XLSX annotation sheet,
then computes extraction precision from the filled sheet.

Workflow:
    1. sample   — draw a reproducible random sample of facts → XLSX
    2. score    — read the annotated XLSX → precision report (overall,
                  per-predicate, per-extraction-method) + Wilson 95% CI

Annotation labels (column `correct`):
    yes      — triple is factually correct w.r.t. the source paper
    no       — triple is wrong (wrong entities, wrong direction, hallucinated)
    partial  — entities right but predicate imprecise (counted as incorrect
               in strict precision, reported separately)

Usage:
    cd backend
    python experiments/annotate_facts.py sample \
        --results experiments/results/experiment_full_llama3.2_latest.json \
        --n 50
    # ... annotate experiments/results/fact_annotation.xlsx ...
    python experiments/annotate_facts.py score \
        --sheet experiments/results/fact_annotation_filled.xlsx
"""

import argparse
import json
import math
import random
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RESULTS_DIR = Path(__file__).resolve().parent / "results"
DEFAULT_SHEET = RESULTS_DIR / "fact_annotation.xlsx"
SEED = 42  # reproducible sampling

COLUMNS = [
    ("fact_id", 12),
    ("source_paper", 28),
    ("subject", 32),
    ("subject_type", 12),
    ("predicate", 20),
    ("object", 32),
    ("object_type", 12),
    ("confidence", 11),
    ("extraction_method", 16),
    ("evidence", 60),
    ("correct", 10),       # expert input: yes / no / partial
    ("notes", 35),         # expert input
]


def cmd_sample(args):
    results_path = Path(args.results)
    if not results_path.exists():
        print(f"ERROR: results file not found: {results_path}")
        sys.exit(1)

    data = json.loads(results_path.read_text())
    facts = data.get("phase2_fact_extraction", {}).get("all_facts", [])
    if not facts:
        print("ERROR: no 'all_facts' in results file. Re-run the experiment "
              "with the updated runner (full/no-rule-engine mode).")
        sys.exit(1)

    n = min(args.n, len(facts))
    rng = random.Random(SEED)
    sample = rng.sample(facts, n)

    wb = Workbook()
    ws = wb.active
    ws.title = "Anotasi"
    header = [c for c, _ in COLUMNS]
    ws.append(header)

    sys_fill = PatternFill("solid", fgColor="DDEBF7")
    exp_fill = PatternFill("solid", fgColor="E2EFDA")
    for idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.fill = exp_fill if name in ("correct", "notes") else sys_fill
        ws.column_dimensions[get_column_letter(idx)].width = width

    for fact in sample:
        ws.append([
            fact.get("fact_id", ""),
            fact.get("source_paper", ""),
            fact.get("subject", ""),
            fact.get("subject_type", ""),
            fact.get("predicate", ""),
            fact.get("object", ""),
            fact.get("object_type", ""),
            fact.get("confidence", 0),
            fact.get("extraction_method", ""),
            fact.get("evidence", ""),
            "", "",
        ])

    dv = DataValidation(type="list", formula1='"yes,no,partial"', allow_blank=True)
    ws.add_data_validation(dv)
    correct_col = get_column_letter(header.index("correct") + 1)
    dv.add(f"{correct_col}2:{correct_col}{n + 1}")
    ws.freeze_panes = "A2"

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Annotation sheet : {out}")
    print(f"Sampled facts    : {n}/{len(facts)} (seed={SEED})")
    print("Label kolom 'correct' dengan yes/no/partial, lalu jalankan subcommand 'score'.")


def wilson_ci(successes: int, n: int, z: float = 1.96):
    """Wilson score 95% confidence interval for a proportion."""
    if n == 0:
        return (0.0, 0.0)
    p = successes / n
    denom = 1 + z**2 / n
    centre = (p + z**2 / (2 * n)) / denom
    margin = z * math.sqrt(p * (1 - p) / n + z**2 / (4 * n**2)) / denom
    return (max(0.0, centre - margin), min(1.0, centre + margin))


def cmd_score(args):
    sheet_path = Path(args.sheet)
    if not sheet_path.exists():
        print(f"ERROR: sheet not found: {sheet_path}")
        sys.exit(1)

    wb = load_workbook(sheet_path, data_only=True)
    ws = wb.active
    header = [str(c.value) for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[idx["fact_id"]] in (None, ""):
            continue
        label = str(row[idx["correct"]] or "").strip().lower()
        if label not in ("yes", "no", "partial"):
            continue
        rows.append({
            "predicate": str(row[idx["predicate"]]),
            "method": str(row[idx["extraction_method"]]),
            "label": label,
        })

    n = len(rows)
    if n == 0:
        print("ERROR: no annotated rows found (kolom 'correct' kosong).")
        sys.exit(1)

    correct = sum(1 for r in rows if r["label"] == "yes")
    partial = sum(1 for r in rows if r["label"] == "partial")
    strict_precision = correct / n
    lenient_precision = (correct + partial) / n
    lo, hi = wilson_ci(correct, n)

    by_predicate = defaultdict(lambda: [0, 0])
    by_method = defaultdict(lambda: [0, 0])
    for r in rows:
        by_predicate[r["predicate"]][1] += 1
        by_method[r["method"]][1] += 1
        if r["label"] == "yes":
            by_predicate[r["predicate"]][0] += 1
            by_method[r["method"]][0] += 1

    report = {
        "n_annotated": n,
        "correct": correct,
        "partial": partial,
        "incorrect": n - correct - partial,
        "strict_precision": round(strict_precision, 3),
        "strict_precision_wilson95": [round(lo, 3), round(hi, 3)],
        "lenient_precision": round(lenient_precision, 3),
        "per_predicate": {
            k: {"correct": c, "total": t, "precision": round(c / t, 3)}
            for k, (c, t) in sorted(by_predicate.items())
        },
        "per_method": {
            k: {"correct": c, "total": t, "precision": round(c / t, 3)}
            for k, (c, t) in sorted(by_method.items())
        },
    }

    out = sheet_path.parent / "fact_precision.json"
    out.write_text(json.dumps(report, indent=2))
    print(f"Report saved: {out}\n")
    print("=== TABEL UNTUK BAB IV ===\n")
    print("| Metrik | Nilai |")
    print("|---|---|")
    print(f"| Fakta dianotasi | {n} |")
    print(f"| Presisi ketat (strict) | {strict_precision:.1%} (95% CI {lo:.1%}–{hi:.1%}) |")
    print(f"| Presisi longgar (dgn partial) | {lenient_precision:.1%} |")
    print()
    print("| Predikat | Presisi | n |")
    print("|---|---|---|")
    for k, v in report["per_predicate"].items():
        print(f"| {k} | {v['precision']:.1%} | {v['total']} |")


def main():
    parser = argparse.ArgumentParser(description="SPO fact precision annotation")
    sub = parser.add_subparsers(dest="command", required=True)

    p_sample = sub.add_parser("sample", help="Sample facts into an XLSX annotation sheet")
    p_sample.add_argument("--results", required=True, help="Experiment result JSON")
    p_sample.add_argument("--n", type=int, default=50, help="Sample size (default 50)")
    p_sample.add_argument("--output", default=str(DEFAULT_SHEET))
    p_sample.set_defaults(func=cmd_sample)

    p_score = sub.add_parser("score", help="Compute precision from annotated sheet")
    p_score.add_argument("--sheet", required=True, help="Filled annotation XLSX")
    p_score.set_defaults(func=cmd_score)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
